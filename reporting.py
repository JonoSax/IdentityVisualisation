import os
import random
from hashlib import sha256
from time import localtime, strftime

import numpy as np
import pandas as pd
from openpyxl import Workbook

from utilities import *

pd.options.mode.chained_assignment = None

"""
TODO:

    - For the outlier detection, tbh straight line distance may not be the best metric for measuring error. A data point may be an outlier because it is out by a lot in one dimension but this might not be captured by a straight line distance.
        UPDATE straight line is indeed the best way because being out on one dimension is arbituary given the orientation of the axis

    - Add descriptions of each input 
"""


class Metrics:

    """
    Object to store and calculate metrics

    The overall goal is to make this as modular as possible so that new reports can be created
    easily. ATM it is kind of bespoke but as new report types are created, parts of these function
    can be easily split off and set as seperate functions for reuse.
    """

    def __init__(
        self,
        plottedIdentities,
        permissions,
        uiddf,
        attribute=None,
        specificTime=None,
        sliderRound=None,
        sliderCluster=None,
        errorTol=1,
    ):

        self.key = uiddf
        self.attribute = attribute
        self.errorTol = errorTol
        self.plottedIdentities = self.setIDdf(plottedIdentities)
        self.permissions = self.setPermdf(permissions)
        self.specificTime = self.getTime(self.plottedIdentities, specificTime)
        self.sliderRound = np.float32(sliderRound)
        self.sliderCluster = np.float32(sliderCluster)

        self.identities = self.plottedIdentities.set_index([self.key, "_DateTime"])

        self.permissions = self.getPermissions(self.plottedIdentities, permissions)

        self.outlierdf = None
        self.dfDistances = None
        self.permissionFixes = None

    def getTime(self, df, val="max"):

        """
        For a given dataframe, get the latest time
        """

        if val == "max":
            dt = df.reset_index()["_DateTime"].max()
        elif val == "min":
            dt = df.reset_index()["_DateTime"].min()
        elif type(val) == int:
            dt = sorted(df.reset_index()["_DateTime"].unique())[val]
        else:
            dt = None

        return dt

    def getPermissions(self, iddf, permdf):

        """
        From the identity dataframe which contains all the identities of interest, get
        the permissiosn which correspond to these identities
        """

        """NOTE issue is that the _DateTime info is a string not an int..."""
        if self.specificTime is not None:
            permbool = (permdf.index.isin(iddf[self.key].unique(), 0)) & (
                permdf.index.isin([self.specificTime], 1)
            )
        else:
            permbool = permdf.index.isin(iddf[self.key].unique(), 0)

        return permdf[permbool]

    def setIDdf(self, df):

        """
        Format the df to be useful for calculations
        """

        # force the id key into a string
        df[self.key] = df[self.key].astype(str)

        # set the multidmiensional positions as floats if they are present
        for d in [dim for dim in df.columns if "Dim" in dim]:
            df[d] = df[d].astype(float)

        # set the datetime value to int
        if "_DateTime" in df.columns:
            df[["_DateTime"]] = df[["_DateTime"]].astype(int)

        # remove spacing from the dataframe column names
        # formatColumns = [r.replace(" ", "") for r in list(df.columns)]
        # df = df.set_axis(formatColumns, axis=1, inplace=False)

        return df

    def setPermdf(self, df):

        """
        Transform the permission dataframe necessary for calculations
        """

        df.set_index(((str(id), int(dt)) for id, dt in df.index), inplace=True)

        return df

    def calculateDistances(self, df=None, attr=None, specificTime=None, sumType="net"):

        """
        Calculate the distances of all object in each attribute unless specified

        Use only the most recent permission information to calculate the centre point, however
        calculate the distances retrospectively to assess the movement of individuals over time
        """

        ret = type(df) == pd.DataFrame
        if df is None:
            df = self.identities.reset_index()

        if attr is None:
            attr = list(df.columns)
            attr = [
                a
                for a in attr
                if "Dim" not in a
                and "_DateTime" not in a  # remove the positions
                and "Unnamed" not in a  # remove the date times
                and self.key not in a  # remove the unnamed column that appears randomly
                and len(df[a].unique())  # remove the unique identifier
                < len(df) / len(df["_DateTime"].unique())
            ]  # remove any entry which appears only once
        elif type(attr) is str:
            attr = [attr]

        self.attribute = attr

        dfAll = df.copy()

        # only calculate distances for the specific time period selected (if provided)
        if specificTime is not None:
            dfAll = dfAll[dfAll["_DateTime"] == specificTime]

        newCols = [f"__Distance{a}" for a in attr]
        for a, newCol in zip(attr, newCols):

            if sumType == "axis":
                newCol = [f"{d}_{newCol}" for d in ["__Dim0", "__Dim1", "__Dim2"]]

            attributeValues = df[a].unique()

            # Calculate the distance of all identities in each element from the median point at the most recently recorded permission extract
            dfCopy = None
            for av in attributeValues:
                dfAttr = df[df[a] == av][
                    ["__Dim0", "__Dim1", "__Dim2", self.key, "_DateTime", a]
                ]

                # NOTE the getTime ALWAYS takes the most recent time information so that the comparsions are made with the most recent data
                centrePoint = np.median(
                    dfAttr[dfAttr["_DateTime"] == self.getTime(dfAttr)][
                        ["__Dim0", "__Dim1", "__Dim2"]
                    ],
                    0,
                )

                # if the error is a net sum distance (ie straight line from the median position of all
                # permission)
                if sumType == "net":
                    dists = np.sum(
                        (dfAttr[["__Dim0", "__Dim1", "__Dim2"]] - centrePoint) ** 2, 1
                    )

                # if the error is calculated per axis
                elif sumType == "axis":
                    dists = np.abs(dfAttr[["__Dim0", "__Dim1", "__Dim2"]] - centrePoint)

                dfAttr[newCol] = dists

                if dfAll is None:
                    dfCopy = dfAttr
                else:
                    dfCopy = pd.concat([dfCopy, dfAttr])

            # add the new column in with the info
            if type(newCol) is not list:
                newCol = [newCol]
            dfAll = pd.merge(
                dfAll,
                dfCopy[[self.key, "_DateTime", *newCol]],
                on=[self.key, "_DateTime"],
                validate="many_to_one",
            )

        dfAll = dfAll.sort_values([self.key, "_DateTime"]).set_index(
            [self.key, "_DateTime"]
        )

        if ret:
            return dfAll
        else:
            self.dfDistances = dfAll

    def findOutliers(self):

        """
        From the distances calculated within each attribute, find identities which are
        outliers (3std away from the median) relative to the other attribute values.

        These identities are then associated with the attribute in which they are an
        outlier
        """

        attrDist = [d for d in self.dfDistances.columns if "__Distance" in d]

        outlierdf = []
        for attr in attrDist:
            outType = attr.split("__Distance")[-1]

            groupAttrDesc = self.dfDistances.groupby(outType)[attr].describe()

            for ele in groupAttrDesc.index:

                # if there are less than 5 identities (arbituary number) then don't include them in outlier calculations because the statistics will be meaningless
                if groupAttrDesc.loc[ele]["count"] < 5:
                    continue

                mean, std = groupAttrDesc.loc[ele][["mean", "std"]]
                q1, q3 = groupAttrDesc.loc[ele][["25%", "75%"]]
                iqr = q3 - q1

                # 3 standard deviations from the mean of distances
                outliers = self.dfDistances[
                    self.dfDistances[attr] > (mean + std * 3 * self.errorTol)
                ]

                # 1.5 x iqr beyond the q3
                """
                outliers = self.dfDistances[
                    (self.dfDistances[attr] > (iqr * 1.5 * self.errorTol + q3))
                    & (self.dfDistances[outType] == ele)
                ]
                """

                outliers["type"] = outType
                if type(outlierdf) is not pd.DataFrame:
                    outlierdf = outliers
                else:
                    outlierdf = pd.concat([outlierdf, outliers])

        self.outlierdf = outlierdf  # .reset_index()

    def outlierEntitlements(self, times=1):

        """
        Find the % likelihood of entitlements contributing to the identities uniqueness to that attribute

        times is the number of time periods to investigate from the most recent permission export
        (ie if times = 1 then only report on the latest time interval based on the outlier information calculated.
        If times = 3 then report on the 3 most recent time intervals where there is outlier information based on
        the median positions of the permissions at the most recent time avaialbe)

        NOTE somewhere in here the privileged permissions need to be excluded and possibly
        a list of exceptions should be imported to prevent re-occuring warnings
        """

        if self.specificTime is None:
            times = sorted(
                self.outlierdf.index.get_level_values(1).unique(), reverse=True
            )[:times]
        else:
            times = [self.specificTime]

        info = None
        for t in times:

            outlierIdsTime = self.outlierdf.loc[(slice(None), t), :].droplevel(1)
            identityTimes = self.identities.loc[(slice(None), t), :].droplevel(1)
            permissionTimes = self.permissions.loc[(slice(None), t), :].droplevel(1).T
            permissionTimes.columns = [p.split("_")[0] for p in permissionTimes.columns]

            for id, data in outlierIdsTime.iterrows():
                idtype = data["type"]

                idattr = self.identities.loc[(id, t)][data["type"]]
                attrIdentities = [
                    i
                    for i in identityTimes[(identityTimes[idtype] == idattr)].index
                    if i in permissionTimes.columns
                ]

                # get the permission to compare and the permissions of the target identity
                relevantPermissions = permissionTimes[attrIdentities]
                targetPermissions = relevantPermissions.pop(id)

                """
                calculate the relative occurence of the entitlements for this attribute distribution:

                 - where diff >0 then the identity has a acquired a permission which is unusual relative to the rest 
                of the attribute population, where 1 means it is ONLY found on that identity
                 - where diff <0 then the identity is missing a permission which is unusual relative to the rest 
                of the attribute population, where -1 means it is ONLY missing for that identity
                """

                diff = (
                    targetPermissions
                    - relevantPermissions.sum(1) / relevantPermissions.shape[1]
                )

                # Using a cut of off 80% (either add or remove) because:
                #   1 - If I included all values then the data frame would become unweildly
                #   2 - 80% is just a nice number, coming from the 80/20 rule
                targetValues = (
                    diff[(diff > 0.8) | (diff < -0.8)]
                    .to_frame("Occurence")
                    .reset_index()
                    .rename(columns={"index": "Value"})
                )

                targetValues[[self.key, "UnixTime", "Attribute", "Element"]] = [
                    id,
                    t,
                    idtype,
                    idattr,
                ]

                if info is None:
                    info = targetValues
                else:
                    info = pd.concat([info, targetValues])

        self.rawOutlierInfo = info

    def getClusterInfo(self, aggDict):

        """
        Get information about the clusters as specified by the slider-round and slider-cluster values
        """

        dfMod = self.plottedIdentities[
            self.plottedIdentities["_DateTime"] == self.specificTime
        ].reset_index(drop=True)
        dfPos = cluster_elements(
            dfMod, self.key, self.attribute, self.sliderRound, aggDict
        )
        dfPos = dfPos.set_index("ClusterID")

        # set the object attributes
        self.dfClusters, _ = getClusterLimit(dfPos, self.attribute, self.sliderCluster)
        self.clusterNames = sorted(self.dfClusters.index)
        self.clusterCount = [int(dfPos.loc[c]["_Count"]) for c in self.clusterNames]

    def getAggregateClusterInfo(self):

        # get the breakdown of the permissions in each cluster
        dfAggregate = pd.DataFrame(None, columns=self.clusterNames)
        for clid, dfCluster in self.dfClusters.iterrows():
            uids = dfCluster[self.key]
            if type(uids) is str:
                uids = [uids]

            dfAggregate[clid] = self.permissions.loc[
                [(i, self.specificTime) for i in uids]
            ].sum(0) / len(uids)

        # sort the aggregate plotIDdf by the order of the number of identities which have the permissions
        # (highest to lowest)
        permissionOrder = (
            self.permissions.loc[(slice(None), self.specificTime), :]
            .sum(0)
            .sort_values(ascending=False)
        )
        dfAggregate = dfAggregate.reindex(permissionOrder.index)

        # remove entries where the value is 0 across all identities (ie no identity has this permission)
        # however create a floor value of 0.01 if there at least one permission for an identity
        dfAggregate[dfAggregate == 0] = None
        self.clusterAgg = np.clip(dfAggregate, 0.01, np.inf).dropna(how="all").fillna(0)


def report_1(
    plotIDdf,
    permissions,
    privData,
    uiddf,
    sliderDate,
    reportName,
    attribute=None,
    errorTol=1,
):

    """

    Create a report with actional information for each permission for each relevant identity
    and provide reasoning for the action. Compare each identities position in relation to other identities with similar attribute information. If their position has changed by some significant amount, include them in a report for possible
    over/under provisioning

    Sheets:
        - Summary,              Some key summary stats
        - PriorityActions,      contains up to 10 tasks which would improve the permission environment
                                modelling is also provided to show the impact of taking these actions
        - Important,            up to 100 additional tasks which could improve the permission environment
        - AllInformation,       contains all the detailed information which was used for the calculation the previous sheets
        - Reference,            notes to explain what information is being provided

    NOTE is a sunburst chart useful? https://plotly.com/python/sunburst-charts/

    NOTE see https://dash.plotly.com/background-callbacks for how to run a callback in the background with visual ques and enable the cancellation of callbacks if it takes too long, progress bar etc

    """

    # ----------- Get base information from the data ----------

    dist = Metrics(
        plotIDdf, permissions, uiddf, specificTime=sliderDate, errorTol=errorTol
    )
    dist.calculateDistances(
        attr=attribute, specificTime=dist.specificTime, sumType="net"
    )
    dist.findOutliers()
    if len(dist.outlierdf) == 0:
        return "No outliers detected"

    dist.outlierEntitlements()

    wb = Workbook()
    del wb["Sheet"]
    # excelExport = pd.ExcelWriter(reportPath)

    # https://github.com/Khan/openpyxl/blob/master/doc/source/tutorial.rst

    # -------------- Summary ----------------
    wsSummary = wb.create_sheet("Summary")
    wsSummary.cell(row=1, column=1).value = "Summary statistics sheet"
    wsSummary.cell(
        row=2, column=1
    ).value = "This sheet provides some high level summary information about your permission environment and actions to take"

    # ----- Summary stats on permissions and attributes/elements -----

    """
    For the latest time period available, for each element group calculate the key statistics 
    regarding the number of permissions, number of identities and variance in permissivity 
    """

    latestIdentities = dist.identities.loc[
        (slice(None), dist.specificTime), :
    ].droplevel(1)
    latestPermissions = dist.permissions.loc[
        (slice(None), dist.specificTime), :
    ].droplevel(1)
    elementAnalysis = pd.DataFrame(
        None,
        columns=[
            "Attribute",
            "Element",
            "IDCount",
            "MinPermissions",
            "LowerQuartile",
            "Median",
            "UpperQuartile",
            "MaxPermissions",
        ],
    )
    elementSpread = pd.DataFrame(
        None,
        columns=[
            "Attribute",
            "Element",
            "IDCount",
            "RelativeStandardDeviation",
            "RelativeSpread",
        ],
    )
    for attr in sorted(dist.attribute):
        element = latestIdentities[attr].unique()

        for ele in sorted(element):
            # get the identities in the specific element from the permissions dataframe
            # DONT merge the data though just because we don't actually need a dataframe that big...
            ids = [
                id
                for id in latestIdentities[
                    latestIdentities[attr] == ele
                ].index.get_level_values(0)
                if id in latestPermissions.index
            ]

            # Analyse the association of permissions per element
            if len(ids) > 0:
                elePermissions = latestPermissions.loc[ids, :]
                idSum = elePermissions.sum(1).describe()
                elementAnalysis.loc[len(elementAnalysis)] = [
                    attr,
                    ele,
                    idSum["count"],
                    idSum["min"],
                    idSum["25%"],
                    idSum["50%"],
                    idSum["75%"],
                    idSum["max"],
                ]

            # Analyse the dispersion of permissions per element (using a minimum of 5 identities)
            if len(ids) > 5 and False:

                # eleSpread = dist.dfDistances.loc[(ids, latestTime), :][f"_Distance{attr}"]
                eleSpread = dist.dfDistances[
                    dist.dfDistances.index.isin(ids, level=0)
                    & dist.dfDistances.index.isin([dist.specificTime], level=1)
                ]
                if len(eleSpread) == 0:
                    continue

                eleSpread = eleSpread[f"_Distance{attr}"]
                idSpread = eleSpread.describe()
                elementSpread.loc[len(elementSpread)] = [
                    attr,
                    ele,
                    idSum["count"],
                    idSpread["std"],
                    idSpread["max"] - idSpread["min"],
                ]

        # ----- Most highly provisioned element groups -----

    """
    Using the median number of permissions per identity in any given element
    """
    rowNo = np.array(
        4
    )  # make a mutable object so the addToReport can iterate on the same variable
    wsSummary.cell(
        row=int(rowNo), column=1
    ).value = "10 most highly provisioned element groups"
    rowNo += 1
    addToReport(wsSummary, elementAnalysis.columns, rowNo)
    elementAnalysis = elementAnalysis.sort_values(
        ["Median", "Element"], ascending=[False, True]
    )
    for _, info in elementAnalysis[:10].iterrows():
        addToReport(wsSummary, info, rowNo)

        # ----- Largest element groups -----
    """
    Using the number of identities per element
    """
    rowNo += 2
    wsSummary.cell(row=int(rowNo), column=1).value = "10 most populated elements"
    rowNo += 1
    addToReport(wsSummary, elementAnalysis.columns, rowNo)
    elementAnalysis = elementAnalysis.sort_values(
        ["IDCount", "Element"], ascending=[False, True]
    )
    for _, info in elementAnalysis[:10].iterrows():
        addToReport(wsSummary, info, rowNo)

        # ----- Greatest permission spread in element groups -----
    """
    Using the normalised standard deviation from the median position of the entitlement per element
    NOTE std not spread becuase spread is closely correlatd with the number of ids vs the actual variance
    in permission position
    """
    rowNo += 2
    wsSummary.cell(
        row=int(rowNo), column=1
    ).value = "10 most disperse permission element groups"
    rowNo += 1
    elementSpread["RelativeStandardDeviation"] /= elementSpread[
        "RelativeStandardDeviation"
    ].max()
    elementSpread["RelativeSpread"] /= elementSpread["RelativeSpread"].max()
    elementSpread = elementSpread.sort_values(
        ["RelativeStandardDeviation", "IDCount"], ascending=[False, False]
    )
    addToReport(wsSummary, elementSpread.columns, rowNo)
    for _, info in elementSpread[:10].iterrows():
        addToReport(wsSummary, info, rowNo)

    # -------------- Priority Actions --------------
    """
    Taking into account both the number of times an individual permission appears as an important
    permission as well as the relative importance to any given identity, select up to 10 permissions
    which should be added to an identity or taken away to improve their relative identity position. 

    NetOccurence:   This value is the absolute sum of the occurences across all identities and the 
                        attributes it affects.
    
    Occurence:      This is the occurence relative to others in the equivalent element defined as:
                        1       = the identity is the only one in the element with the permission 
                                    (unique exception)
                        1>      = the identity has the permission which is unusual 
                                    (rare exception)
                        1>>     = the identity DOES has a permission which most other identities in the element  
                                    also have 
                                    (common exception/other identities under-privileged)
                        0       = the permission is found in all identities 
                                    (element wide permission)
                        -1<<    = the identity DOES NOT have a permission which most other identities in
                                    the element also do not have 
                                    (uncommon exception/other identities over-privileged)
                        -1<     = the identity does not have a permission which is unusual 
                                    (rare exculision )
                        -1      = this identity is the only one in that element without the permission 
                                    (unique exculsion)


                    
    """

    wsPriorityActions = wb.create_sheet("Priority Actions")
    wsPriorityActions.cell(row=1, column=1).value = "Priority actions to resolve"
    wsPriorityActions.cell(
        row=2, column=1
    ).value = "The following information outlines the 10 permissions which are causing the greatest identity discrepancy"

    # ------ Calculations to specify the impact of permission modification ------
    priority = dist.rawOutlierInfo.copy()
    priority["NetOccurence"] = np.abs(priority["Occurence"])
    priority = (
        priority.groupby("Value")
        .sum("NetOccurence")
        .sort_values("NetOccurence", ascending=False)
        .reset_index()
    )

    priorityPermissions = list(priority["Value"]).copy()

    rowNo = np.array(4)
    removeC = 0
    addC = 0
    removePermission = {}
    addPermission = {}
    idsImpacted = []
    priorityInfo = []
    # For up to 10 permissions (max of 7 to either add or remove), investigate the permissions to
    # prioritise actions for
    n = 0
    while (
        removeC + addC < 10
        and len(priorityPermissions) > 0
        and n < len(priorityPermissions)
    ):
        pi = priorityPermissions[n]
        if pi in privData.index:
            pi = f"{pi} **Privilege level {int(privData.loc[pi])}"

        pAnalysis = dist.rawOutlierInfo[dist.rawOutlierInfo["Value"] == pi]

        # To remove the permissions
        removeP = pAnalysis[pAnalysis["Occurence"] > 0]
        if len(removeP) > 0 and removeC < 7:
            ids = list(removeP[dist.key].unique())
            removeinfo = [
                pi,
                "Remove from identity",
                ", ".join(ids),
                ", ".join(list(removeP["Attribute"].unique())),
                f"{int((np.abs(removeP['Occurence']).min())*100)}% - {int((np.abs(removeP['Occurence']).max())*100)}%",
            ]
            priorityInfo.append(removeinfo)
            removeC += 1
            removePermission[pi] = ids
            idsImpacted += ids
            priorityPermissions[n] = None

        # To add the permissions
        addP = pAnalysis[pAnalysis["Occurence"] < 0]
        if len(addP) > 0 and addC < 7:
            ids = list(addP[dist.key].unique())
            addinfo = [
                pi,
                "Add to identity",
                ", ".join(ids),
                ", ".join(list(addP["Attribute"].unique())),
                f"{int((np.abs(addP['Occurence']).min())*100)}% - {int((np.abs(addP['Occurence']).max())*100)}%",
            ]
            priorityInfo.append(addinfo)
            addC += 1
            addPermission[pi] = ids
            idsImpacted += ids
            priorityPermissions[n] = None

        n += 1

    # remove any priority permissions that were used
    priorityPermissions = [p for p in priorityPermissions if p is not None]

    # get the unique entries of ids
    idsImpacted = list(set(idsImpacted))

    """        
    Re-run the MDS with the added and removed permissions and re-calculate the distances of each 
    identity. 

    Identify the % changed distances from the median of their element and whether it is now within 
    the acceptable boundaries or if it requires further action (ie see the important actions sheet). 
    """

    latestPermissions = (
        dist.permissions.loc[(slice(None), dist.specificTime), :].droplevel(1).copy()
    )
    modLatestPermissions = latestPermissions.copy()

    # create the permission data frame from the latest time export
    latestPermissions[[dist.key, "Type"]] = [
        [m.split("_")[0], "Original"] for m in latestPermissions.index
    ]
    latestPermissions = latestPermissions.set_index([dist.key, "Type"])

    # create a permission data frame to modify the permission environment
    modLatestPermissions[[dist.key, "Type"]] = [
        [m.split("_")[0], "Modified"] for m in modLatestPermissions.index
    ]
    modLatestPermissions = modLatestPermissions.set_index([dist.key, "Type"])

    for a in addPermission:
        ids = addPermission[a]
        modLatestPermissions[a].loc[ids] = 1

    for r in removePermission:
        ids = removePermission[r]
        modLatestPermissions[r].loc[ids] = 0

    allPerm = pd.concat([latestPermissions, modLatestPermissions])

    # re-calculate the mds
    perMos = mdsCalculation(allPerm, privData, method="isomap")
    dimNames = [f"__Dim{n}_R" for n in range(3)]

    # merge all the positional and identity data together
    entitleExtract = pd.DataFrame(
        np.hstack([perMos, np.array(list(allPerm.index))]),
        columns=[*dimNames, *allPerm.index.names],
    )
    entitleExtract = entitleExtract.merge(
        dist.identities.loc[(slice(None), dist.specificTime), :].reset_index(),
        on=dist.key,
    )
    entitleExtract[["__Dim0", "__Dim1", "__Dim2"]] = entitleExtract[
        ["__Dim0_R", "__Dim1_R", "__Dim2_R"]
    ].astype(float)
    entitleExtract["_DateTime"] = entitleExtract["_DateTime"].astype(int)
    entitleExtract.loc[entitleExtract["Type"] == "Modified", "_DateTime"] = (
        dist.specificTime - 1
    )  # iterate the latest time to create the "future" permission

    #### RERUN THE OUTLIER CALCULATIONS AND THE CHANGE IN DISTANCE FROM THE MEDIAN POINT #####
    dfNew = dist.calculateDistances(entitleExtract, dist.attribute)
    dfIDsofInterest = dfNew.loc[idsImpacted]
    distAttr = [d for d in dfIDsofInterest.columns if "_Distance" in d]
    oldDists = dfIDsofInterest.loc[(slice(None), dist.specificTime), :][
        distAttr
    ].droplevel(1)
    newDists = dfIDsofInterest.loc[(slice(None), dist.specificTime - 1), :][
        distAttr
    ].droplevel(1)
    distInfo = pd.DataFrame(1 - newDists / oldDists)  # the % reduction in distance
    distInfo = distInfo.sort_index(
        key=distInfo.median(1).get, ascending=False
    )  # sort the distinfo by the median value of all attributes calculated for from largest to smallest
    distDesc = pd.Series(np.hstack(distInfo.to_numpy())).describe()

    # find the identities in each element which are the closest to the median position
    minIds = [
        dist.dfDistances.groupby(ele.replace("__Distance", ""))[distAttr].idxmin()[ele]
        for ele in distAttr
    ]
    bestFitIDs = [
        [
            [id, create_datetime(dt), minIdAttr.index.name, ele]
            for ele, (id, dt) in minIdAttr.iteritems()
        ]
        for minIdAttr in minIds
    ]
    bestFitIDs = [item for sublist in bestFitIDs for item in sublist]

    # ----- Summary statement -----

    prioritySummary = [
        "The impact of performing the following modification to the identities and permissions is a:",
        f"{int(distDesc['min']*100)} - {int(distDesc['max']*100)}% reduction in permission distances across all elements for an median of {int(distDesc['50%']*100)}%",
        (
            f"{distInfo.T.max().idxmax()} from {dist.identities.loc[distInfo.T.max().idxmax()].loc[dist.specificTime][dist.attribute][0]}"
            f" in attribute {distInfo.max().idxmax().replace('__Distance', '')} will improve the most"
        ),
    ]

    for p in prioritySummary:
        addToReport(wsPriorityActions, [p], rowNo)

    rowNo += 1
    addToReport(
        wsPriorityActions,
        ["The identities which are the best representation of the elements are:"],
        rowNo,
    )
    addToReport(
        wsPriorityActions, [uiddf, "Permission date", "Attribute", "Element"], rowNo
    )
    for p in bestFitIDs:
        addToReport(wsPriorityActions, p, rowNo)

    rowNo += 1

    # ----- ID specific changes -----

    addToReport(
        wsPriorityActions, ["Identity", "%Access variation change by attribute"], rowNo
    )
    attrColumns = [d.replace("__Distance", "") for d in distInfo.columns]
    attrColLabel = sorted(attrColumns * 2)
    attrColLabel[0::2] = [""] * len(attrColumns)
    addToReport(
        wsPriorityActions,
        attrColLabel,
        rowNo,
    )
    for id, values in distInfo.iterrows():
        addToReport(
            wsPriorityActions,
            [id]
            + [dist.identities.loc[id, dist.specificTime][ele] for ele in attrColumns]
            + [f"{int(v*100)}%" for v in values.to_list()],
            rowNo,
        )
    rowNo += 1

    # ----- Actions to take -----

    addToReport(
        wsPriorityActions,
        [
            "Permissions to action",
            "Action to take",
            "Identities impacted",
            "Attributes impacted",
            "Likelihood of impact",
        ],
        rowNo,
    )

    # write the priority report
    for info in priorityInfo:
        addToReport(wsPriorityActions, info, rowNo)

    ##### -------------- Important Actions --------------
    wsImportantActions = wb.create_sheet("Important Actions")
    wsImportantActions.cell(
        row=1, column=1
    ).value = "Actions which would improve your permission environment but should be secondary to the Priority actions"
    wsImportantActions.cell(
        row=2, column=1
    ).value = "The following information outlines up to 30 other permissions which are causing the greatest identity discrepancy"

    addToReport(
        wsImportantActions,
        [
            "Permissions to action",
            "Action to take",
            "Identities impacted",
            "Attributes impacted",
            "Likelihood of impact",
        ],
        4,
    )

    removeC = 0
    addC = 0
    rowNo = 5
    while removeC + addC < 30 and len(priorityPermissions) > 0:
        pi = priorityPermissions.pop(0)
        if pi in "Access0" in privData.index:
            pi = f"{pi} **Privilege level {int(privData.loc[pi])}"

        pAnalysis = dist.rawOutlierInfo[dist.rawOutlierInfo["Value"] == pi]

        # To remove the permissions
        removeP = pAnalysis[pAnalysis["Occurence"] > 0]
        if len(removeP) > 0 and removeC < 20:
            removeinfo = [
                pi,
                "Remove from identity",
                ", ".join(list(removeP[dist.key].unique())),
                ", ".join(list(removeP["Attribute"].unique())),
                f"{int((np.abs(removeP['Occurence']).min())*100)}% - {int((np.abs(removeP['Occurence']).max())*100)}%",
            ]
            addToReport(wsImportantActions, removeinfo, rowNo)
            removeC += 1
            rowNo += 1

        # To add the permissions
        addP = pAnalysis[pAnalysis["Occurence"] < 0]
        if len(addP) > 0 and addC < 20:
            addinfo = [
                pi,
                "Add to identity",
                ", ".join(list(addP[dist.key].unique())),
                ", ".join(list(addP["Attribute"].unique())),
                f"{int((np.abs(addP['Occurence']).min())*100)}% - {int((np.abs(addP['Occurence']).max())*100)}%",
            ]
            addToReport(wsImportantActions, addinfo, rowNo)
            addC += 1
            rowNo += 1

    ##### -------------- All info ------------------
    wsAllInformation = wb.create_sheet("Raw Outlier Data")
    wsAllInformation.cell(
        row=1, column=1
    ).value = "Raw data used for the Priority and Important actions"
    rowNo = np.array(2)
    addToReport(wsAllInformation, list(dist.rawOutlierInfo.columns), rowNo)

    """
    for r in dataframe_to_rows(dist.rawOutlierInfo, index=True, header=True):
        wsAllInformation.append(r)
    """
    for _, r in dist.rawOutlierInfo.iterrows():
        addToReport(wsAllInformation, list(r), rowNo)

    # -------------- Reference documentation --------------
    wsReference = wb.create_sheet("Reference")
    wsReference.cell(
        row=1, column=1
    ).value = "This sheet has general information which explains the results and the data that has been used to calculate the summaries."

    ##### ---------- Explaining the priority actions sheet ----------

    wsReference.cell(row=3, column=1).value = "Priority Actions sheet:"
    wsReference.cell(
        row=3, column=2
    ).value = "This workbook describes the specific actions which, if taken, have been modelled to \nalter the permission environment. This will MOSTLY positive improvements in the permission \nspace (ie will reduce identity spread and increase clusterting). However negative \nimprovements can occur, most often when there are sub-clusters of the modelled element \nwhich makes it hard for the identity to be optimised. This does not invalidate the entire \nresults, moreso it just requires qualitative analysis as to whether there are sub-clusters. \nIdeally these sub-clusters would be seperated somehow into distinct clusters."

    wsReference.cell(row=4, column=2).value = "Permissions to action"
    wsReference.cell(
        row=4, column=3
    ).value = "The permission which the action needs to be applied to"
    wsReference.cell(row=5, column=2).value = "Action to take"
    wsReference.cell(
        row=5, column=3
    ).value = (
        "Whether the permission needs to be added or removed from the identities listed"
    )
    wsReference.cell(row=6, column=2).value = "Identities impacted"
    wsReference.cell(
        row=6, column=3
    ).value = "List of identities which have been identified as either needed the permission added or removed"
    wsReference.cell(row=7, column=2).value = "Attributes impacted"
    wsReference.cell(
        row=7, column=3
    ).value = "The specific attribute in which the modelling has shown the improvement. This \nis included because this report can model the impact of permission changes across all \nattributes in a given set of identities, however its msot common use case will be for \nspecific attribute anlaysis."
    wsReference.cell(row=8, column=2).value = "Likelihood of impact"
    wsReference.cell(
        row=8, column=3
    ).value = "The % contributing factor that the specific attribute and action taken on the \nspecified identity contributed to the change in identity permissivity. This is from \n0-100% where 0% indicates the action made no difference and 100% means it definitely \nmade a difference. Values between these indicate the relative impact that any given \naction like had on the results of the re-modelling. This is calculated by assessing the \n% of other identities who either have or don't have the permission. An identity which \nhas a unique permission that is then removed, and conversly an identity who is the only \none without a permission which then has it added, will have a 100% chance of reducing \nthe permissivity distance because this unique permission will increase the identities \ndifference relative to all other identities. If an permission is not unique to an \nidentity then there is a chance that adding/removing the permission will actually \nincrease the permissivity distance for a small number \n identities who were similar to \nthat identity beforehand. This is because of the a complex relationships between all \nidentities which are simplified with the MDS algorithm which serves to minimise the \nglobal errors sometimes at the expense of local errors. It is impossible to individually \nmodel the impact of every permission in isolation because by definition the permissivity \ndistances is a results of the relationship between identities NOT in isolation. For the \nmodelling, only permissions which will have a contribution score of at least 80% are \nconsidered."

    # create the report as an excel and save in the downloads
    timeInfo = strftime("%Y-%m-%d_%H.%M.%S", localtime())
    name = f"{reportName}_{timeInfo}"
    reportPath = f"{os.path.expanduser('~')}/Downloads/{name}.xlsx"
    print(f"----- {reportPath} created -----")
    wb.save(filename=reportPath)

    return f"{name} saved"


def report_2(
    plotIDdf,
    permissions,
    uiddf,
    attribute,
    sliderRound,
    sliderCluster,
    sliderDate,
    hover_data,
    reportName,
):

    """
    Report on the clusters that are currently observable in the data based on the value of the sliders.
    Report on the permissions and the other identity attributes and their % breakdown for
    all identities within each partcular bubble.
    """

    wb = Workbook()
    del wb["Sheet"]
    # excelExport = pd.ExcelWriter(reportPath)

    # https://github.com/Khan/openpyxl/blob/master/doc/source/tutorial.rst

    # ---------- Get the base objects and dataframes ----------

    # create the dictionary aggregation rule
    aggDict = {hd: lambda x: list(x) for hd in hover_data if hd != attribute}

    # get the clustering information
    pos = Metrics(
        plotIDdf, permissions, uiddf, attribute, sliderDate, sliderRound, sliderCluster
    )
    pos.getClusterInfo(aggDict)
    pos.getAggregateClusterInfo()

    # ---------- Identify the breakdown of permissions in each cluster ----------

    """
    Report on the permissions of each cluster and the proporption of identities which have it in each cluster
    """

    wsPermAnalysis = wb.create_sheet("Permission analysis")
    wsPermAnalysis.cell(
        row=1, column=1
    ).value = "Analysis of the permission per cluster"
    wsPermAnalysis.cell(
        row=2, column=1
    ).value = "The following information breaks down the % of identities each cluster who have the corresponding permission."
    wsPermAnalysis.cell(
        row=3, column=1
    ).value = "The permissions are ordered from the most to least common across all individuals (including those not clustered)."
    rowNo = np.array(5)

    addToReport(wsPermAnalysis, ["Cluster ID"] + pos.clusterNames, rowNo)
    # NOTE using clustername because the cluster names are assigned based on the number of identities present (ie 0 is the largest cluster, 1 is the next largest etc)
    addToReport(
        wsPermAnalysis,
        ["Element clustering"]
        + [pos.dfClusters.loc[clID][attribute] for clID in pos.clusterNames],
        rowNo,
    )
    addToReport(wsPermAnalysis, ["Cluster Count"] + pos.clusterCount, rowNo)
    addToReport(wsPermAnalysis, ["Permission Value"], rowNo)
    rowNo += 1
    for idx, dfag in pos.clusterAgg.iterrows():
        addToReport(wsPermAnalysis, [idx] + [np.round(d, 2) for d in dfag], rowNo)

    # ---------- Report on the permissiosn which make this cluster differentiated from the other clusters ----------

    """
    It is valuable to understand why the clusters have been created. At a high level this can be explaiend by 
    the fact that different clusters will different sets of common permissions. 

    HOWEVER it is extremely complicated to model the degree which a single permission has influence over the 
    resultant clutsering, especially when identities within a cluster do not all have the same permissions. Also, 
    given the 3D positions are discovered by an iterative machine learning algorithm, we are essentailly trying 
    to understand why the "black box" algorithm has made its "decision" which is notoriously difficult.
    
    One fact is certain though: permissions which are uniquely in only one cluster contribute significantly to  
    seperation of the cluster from other clusters.    
    """

    wsPermUnique = wb.create_sheet("Differentiated permissions")
    wsPermUnique.cell(row=1, column=1).value = "Analysis of the unique permissions"
    wsPermUnique.cell(
        row=2, column=1
    ).value = "The following sheet reports on the unique permissions of each cluster and the fraction of identities in the cluster which have it."
    wsPermUnique.cell(
        row=3, column=1
    ).value = "The unique permissions are listed per cluster"

    dfNet = pos.clusterAgg * pos.clusterCount
    diffDict = {}
    for cN, cC in zip(pos.clusterNames, pos.clusterCount):
        dfExclude = dfNet[[c for c in pos.clusterNames if c != cN]].sum(1) / (
            sum(pos.clusterCount) - cC
        )
        pos.clusterAgg[cN]
        # diff = pos.clusterAgg[cN] - dfExclude
        diff = (
            pos.clusterAgg[cN] / dfExclude
        )  # seach for the np.inf to find the unique permissions?
        # ad = diff.describe()
        # iqr = ad["75%"] - ad["25%"]
        # diffDict[cN] = diff[(diff>iqr*1.5+ad["75%"]) | (diff<ad["25%"]-iqr*1.5)]
        diffDict[cN] = pos.clusterAgg.loc[list(diff[diff == np.inf].index)][
            cN
        ]  # store only the permissions which are unique in each cluster and the % occurence in their cluster

    colNo = np.array(1)
    wsPermUnique.cell(row=5, column=2).value = f"% identities with permission"
    for n, key in enumerate(diffDict.keys()):
        clusterInfo = diffDict[key].sort_values(ascending=False)
        wsPermUnique.cell(row=5, column=(n * 3 + 1)).value = f"{key} permissions"
        addToReport(wsPermUnique, list(clusterInfo.index), colNo=colNo, rowStart=6)
        addToReport(
            wsPermUnique,
            [np.clip(np.round(c, 2), 0.01, np.inf) for c in clusterInfo],
            colNo=colNo,
            rowStart=6,
        )
        colNo += 1

    # ---------- Identify the breakdown of elements of each attribute in each cluster ----------

    """
    Report of the breakdown of elements in each cluster. 

    Each cluster has been formed around a single element but obvioulsy the other attributes do not necessarily
    align so this is a breakdown per cluster of each element %
    """

    wsAttrAnalysis = wb.create_sheet("Attribute analysis")
    wsAttrAnalysis.cell(
        row=1, column=1
    ).value = "Analysis of the attributes and elements per cluster"
    wsAttrAnalysis.cell(
        row=2, column=1
    ).value = "The following information breaks down the % of elements in cluster for each identity attribute."
    wsAttrAnalysis.cell(
        row=3, column=1
    ).value = "Each table is ordered from the one with the least to the most number of distinct elements."
    rowNo = np.array(5)

    # Select the attributes for analysis
    analyseAttrs = [
        h for h in hover_data if h.find("_") == -1 and h != uiddf and h != attribute
    ]

    # get the breakdown of the other attributes in each cluster
    attrBreakDowns = {}
    for attr in analyseAttrs:
        allDfAttr = pos.identities.loc[(slice(None), pos.specificTime), :][
            attr
        ].unique()

        # if the number of elements for this attribute is more than half the number of identities and is more than 100 unique combinations then don't process
        if len(allDfAttr) > len(pos.identities) / 2 and len(allDfAttr) > 100:
            continue
        dfAttr = pd.DataFrame(None, index=sorted(allDfAttr), columns=pos.clusterNames)
        dfCluster = pos.dfClusters[[attr]]
        for clid, dc in dfCluster.iterrows():
            elements = dc[attr]
            if type(elements) is not list:
                elements = [elements]

            # update each element with the proporption of each element in the cluster
            for e in list(set(elements)):
                dfAttr.loc[e].loc[clid] = (
                    np.array(elements) == e
                ).sum() / elements.__len__()

        # drop rows where there are no inputted values and replace nan with 0's. Sort the table from the most identities is the clusters to least
        dfAttr = dfAttr.dropna(how="all").fillna(0)
        dfAttr["_sum"] = (dfAttr * pos.clusterCount).sum(1)
        dfAttr = (
            dfAttr.reset_index()
            .sort_values(by=["_sum", "index"], ascending=[False, True])
            .set_index("index")
            .drop("_sum", axis=1)
        )
        attrBreakDowns[attr] = dfAttr

    # order the attributes by the least number of unique elements first
    attrOrderX = np.array([len(attrBreakDowns[a]) for a in attrBreakDowns]).argsort()
    attrOrder = [list(attrBreakDowns.keys())[x] for x in attrOrderX]

    addToReport(wsAttrAnalysis, ["Cluster ID"] + pos.clusterNames, rowNo)
    addToReport(
        wsAttrAnalysis,
        ["Element Clustering"]
        + [pos.dfClusters.loc[clID][attribute] for clID in pos.clusterNames],
        rowNo,
    )
    addToReport(wsAttrAnalysis, ["Cluster Count"] + pos.clusterCount, rowNo)
    rowNo += 1
    for attr in attrOrder:
        attrDf = attrBreakDowns[attr]
        attrDf = attrDf
        addToReport(wsAttrAnalysis, [f"Attribute: {attr}"], rowNo)
        for idx, dfag in attrDf.iterrows():
            addToReport(
                wsAttrAnalysis,
                [idx]
                + [
                    0 if d == 0 else np.clip(np.round(d, 2), 0.01, np.inf) for d in dfag
                ],
                rowNo,
            )
        rowNo += 1

    timeInfo = strftime("%Y-%m-%d_%H.%M.%S", localtime())
    name = f"{reportName}_{timeInfo}"
    reportPath = f"{os.path.expanduser('~')}/Downloads/{name}.xlsx"
    wb.save(filename=reportPath)

    output = f"{name} saved"

    return output


def report_3(plotIDdf, permissions, uiddf, attribute, hover_data, reportName):

    """
    Report on the trend of individual identity permissions within each element

    Sheet layout:
        - Permission changes
            - Each permission that has been added/removed across all identities
            - The number that have been added/removed
            - The specific identities that have been added/removed
            - The break down of attributes that have had it added/removed

        - Attribute changes
            -

    - Get every unique permission that has been changed and report on the individuals who have got/lost it

    - Get every unique attribute that has been changed and report on what has changed

    """

    select_data = [h for h in hover_data if h.find("DateTime") == -1]

    # add random changes for TESTING purposes
    plotIDdf = add_random_changes(
        plotIDdf,
        cols=[
            "Department",
            "Functional Division",
            "Group Division",
            "Job Profile",
            "Management Level",
        ],
    )

    # plotIDdf.sort_values("_DateTime", inplace=True)
    idAttrRep = pd.DataFrame(None)
    idPermRep = pd.DataFrame(None)
    for id in plotIDdf[uiddf].unique():

        # assess the change in attribute information
        idIDdf = plotIDdf[plotIDdf[uiddf] == id][["_DateTime"] + select_data]
        idAttrRep = pd.concat([monitor_change(idIDdf, id), idAttrRep])

        # assess the change in permission information
        idPermdf = permissions.loc[id].sort_index()
        idMatrix = idPermdf.to_numpy()
        permChanges = idPermdf.loc[:, ~(idMatrix[0] == idMatrix).all(0)].reset_index()
        idPermRep = pd.concat([idPermRep, monitor_change(permChanges, id)])

    # create a chord diagram of the changes in attributes
    attrGroups = (
        idAttrRep.groupby(["ele0", "ele1"])
        .agg({"value": lambda x: list(set(x))[0], "id": "count"})
        .sort_values("id")
    )

    """
    # Creat a Chord diagram 
    # NOTE this is not working....
    eleCats = sorted(
        list(
            set(
                [
                    item
                    for t in attrGroups["value"][attrGroups["value"] == attribute].index
                    for item in t
                ]
            )
        )
    )
    eleLabels = sorted([f"{e}_From" for e in eleCats] + [f"{e}_To" for e in eleCats])
    eleMove = pd.DataFrame(0, index=eleLabels, columns=eleLabels)
    # create the movement matrix
    for eleFrom in eleCats:
        attrTo = attrGroups.loc[eleFrom]

        for eleTo in attrTo.index:
            eleMove.loc[f"{eleFrom}_From", f"{eleTo}_To"] = attrTo.loc[eleTo]["id"]
            eleMove.loc[f"{eleTo}_To", f"{eleFrom}_From"] = attrTo.loc[eleTo]["id"]
    
    a = holoviews.Chord(eleMove.values)
    holoviews.save(a, "out.html")

    # insert the chord diagram into the outputted excel doc
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    img = openpyxl.drawing.image.Image('test.jpg')
    img.anchor = 'A1'
    ws.add_image(img)
    wb.save('out.xlsx')
    """


def save_selected_information(selectIDdf, permdf, uiddf, filePath):

    """
    Save the selected information

    iddf, identity dataframe

    permdf, permission dataframe

    Sheets:
        Identities: the identities and all their attributes
        Permissions: the identities and all their permissions
    """

    wb = Workbook()
    del wb["Sheet"]

    wsIdentities = wb.create_sheet("Info attributes")
    wsPermissions = wb.create_sheet("Permissions")

    # --------------- Permissions sheet ---------------

    wsPermissions.cell(row=1, column=1).value = "Permissiosn of the identity"
    rowNo = np.array(3)

    # seperate the selected data between clusters and individual identities. A cluster contains 2 or more IDs
    # add hash info
    selectIDdf["__hashinfo"] = [
        sha256(info.to_json().encode()).hexdigest()[:8]
        for _, info in selectIDdf.iterrows()
    ]

    selectIDdf.drop_duplicates("__hashinfo", inplace=True)

    clusterRows = np.array([type(id) is list for id in selectIDdf[uiddf]])
    clusterdf = selectIDdf[clusterRows]
    identitydf = selectIDdf[~clusterRows]

    # get all the permissions of the selected individual identities in the order they were selected
    permsID = pd.DataFrame(
        [
            permdf.loc[(id, int(dt)), :]
            for _, (id, dt) in identitydf[[uiddf, "_DateTime"]].iterrows()
            if type(id) is not list
        ]
    ).T

    # get the %proporption permissions for the selected clusters in the order they were selected if there are any
    if len(clusterdf) > 0:

        # If there are clusters then the 2nd to last column will contain the cluster information. The last column is the newly created __hashinfo column
        clusterdf.rename(columns={clusterdf.columns[-2]: "ClusterID"}, inplace=True)
        selectIDdf.rename(columns={clusterdf.columns[-2]: "ClusterID"}, inplace=True)

        permsIDCluster = pd.DataFrame(
            [
                pd.DataFrame([permdf.loc[(id, int(dt)), :] for id in ids]).mean()
                for _, (ids, dt) in clusterdf[[uiddf, "_DateTime"]].iterrows()
            ]
        ).T

        index = [
            (
                f"{info['ClusterID']} ({info['__hashinfo']})",
                info["_DateTime"],
            )
            for _, info in clusterdf.iterrows()
        ]
        permsIDCluster.set_axis(
            labels=pd.MultiIndex.from_tuples(index), axis=1, inplace=True
        )

        permsID = pd.concat(objs=[permsID, permsIDCluster], axis=1)

    permsID = permsID[np.sum(permsID, 1) > 0]

    # sort the permission df rows by the number of permissions (highest to lowest) then alphabetically and (NOT DOING THIS) the columns by the uiddf alphabetically (same as the selectIDdf)
    permsID["_sum"] = permsID.sum(1)
    permsID = (
        permsID.reset_index()
        .sort_values(by=["_sum", "Value"], ascending=[False, True])
        .set_index("Value")
        .drop("_sum", axis=1)
    )
    # permsID.reindex(sorted(permsID.columns), axis=1)

    addToReport(wsPermissions, permsID.columns.get_level_values(0), rowNo, colStart=2)
    addToReport(
        wsPermissions,
        ["Accesses"]
        + [create_datetime(t) for t in permsID.columns.get_level_values(1)],
        rowNo,
    )
    for perm, idInfo in permsID.iterrows():
        addToReport(wsPermissions, [perm] + list(idInfo), rowNo)

    # --------------- Identities attribute info sheet ---------------

    wsIdentities.cell(row=1, column=1).value = "Selected identities/cluster info"
    # Select any attribute names which DONT start with "_"
    selectAttr = [c for c in selectIDdf.columns if c[0] != "_"]

    rowNo = np.array(3)
    addToReport(wsIdentities, selectAttr, rowNo)
    for _, idInfo in selectIDdf.iterrows():
        addToReport(wsIdentities, idInfo[selectAttr], rowNo)

    # --------------- Cluster identities ---------------

    # create cluster identities tab
    if len(clusterdf) > 0:
        # NOTE this is split from the above if statement purely for readability/organisational reasons.
        wsClusterIdentities = wb.create_sheet("Cluster identities")
        wsClusterIdentities.cell(row=1, column=1).value = "Identities in each cluster"

        colNo = np.array(1)
        addToReport(
            wsClusterIdentities,
            [f"{id}: {create_datetime(dt)}" for id, dt in index],
            np.array(3),
        )
        for _, idInfo in clusterdf.iterrows():
            addToReport(
                wsClusterIdentities,
                sorted(idInfo[uiddf]),
                colNo=colNo,
                rowStart=np.array(4),
            )

    # add a new tab to the report and report on the identities in each cluser

    wb.save(filename=filePath)


def monitor_change(dfData, id, dtAttr="_DateTime"):

    """
    From a dataframe with time and the measured element, report on what changed and when

    dfData (pd.DataFarme): datafarme containing the identity data. The first column MUST contain the _DateTime information and all subsequent columns containing the information which is being monitored

    id (str): string which describes the unique identifier for this change monitoring

    dtAttr (str): attribute value which describes the datetime used
    """

    # if the time named attribute is not in the first column then move it
    if dtAttr != dfData.columns[0]:
        dfDataCols = [dtAttr] + list(dfData.columns).remove(dtAttr)
        dfData = dfData[dfDataCols]

    dfStore = pd.DataFrame(None, columns=["id", "value", "ele0", "ele1", "dt0", "dt1"])

    timeAttr, *values = dfData.columns
    dfData.sort_values(timeAttr, inplace=True)
    for value in values:
        # if the value is unique for all time periods, don't loop through
        if len(set(dfData[value])) == 1:
            continue

        # if there are multiple values along the time periods, loop through and identity when and what changed
        cDt, cEle = dfData[[timeAttr, value]].iloc[0]
        for _, (nDt, nEle) in dfData[[timeAttr, value]][1:].iterrows():
            if nEle != cEle:

                # report the value, current element, new element, current datetime of the new element and the datetime which the current datetime was FIRST detected
                dfStore.loc[len(dfStore)] = [id, value, cEle, nEle, cDt, nDt]

                # if the element has changed, update the
                cEle = nEle
                cDt = nDt

    return dfStore


def add_random_changes(df: pd.DataFrame, num=100, cols=None):

    """
    Add random changes to the dataframe by replacing an existing value in a row with a randomally selected value in the same columns

    num = the number of changes to make
    cols = list of columns
    """

    if cols is None:
        colPos = np.arange(len(df.columns))
    else:
        colList = list(df.columns)
        colPos = []
        for c in cols:
            if c in colList:
                colPos.append(list(df.columns).index(c))

    for _ in range(num):
        c = colPos[random.randint(0, len(colPos) - 1)]

        r0 = random.randint(0, len(df) - 1)
        r1 = random.randint(0, len(df) - 1)

        df.iloc[r0, c] = df.iloc[r1, c]

    return df


if __name__ == "__main__":

    df = pd.read_csv(
        "C:/Users/jreshef/Documents/Projects/PermissionAnalysis/results/Identity3D_1659964744.csv",
        dtype=str,
    )

    attribute = "Department"
    uiddf = "Username"

    # report_1(df, permissions, identities, attribute, uiddf)
